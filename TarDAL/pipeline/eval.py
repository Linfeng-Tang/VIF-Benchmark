from pathlib import Path
from typing import List

import torch
import torch.backends.cudnn
from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from utils.image_pair import ImagePair
import time
import numpy as np
import os
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def writexls(save_name, method_name, time_list, sheet_name, i):
    df = pd.DataFrame({method_name: time_list})
    append_df_to_excel(save_name, df, sheet_name=sheet_name, index=False, startrow=0, startcol=i)

#网络参数数量
def get_parameter_number(net):
    total_num = sum(p.numel() for p in net.parameters())
    return total_num


class Eval:
    def __init__(self, net, cudnn: bool = True, half: bool = False, eval: bool = False):
        torch.backends.cudnn.benchmark = cudnn
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.half = half
        _ = net.half() if half else None
        _ = net.to(self.device)
        _ = net.eval() if eval else None
        self.net = net

    @torch.no_grad()
    def __call__(self, ir_paths: List[Path], vi_paths: List[Path], dst: Path, color: bool = False):
        os.makedirs(dst, exist_ok=True)
        file_list = []
        time_list = []
        p_bar = tqdm(enumerate(zip(ir_paths, vi_paths)), total=len(ir_paths))
        for idx, (ir_path, vi_path) in p_bar:
            # print(ir_path, vi_path)
            assert ir_path.stem == vi_path.stem
            p_bar.set_description(f'fusing {ir_path.stem} | device: {str(self.device)}')
            file_list.append(ir_path.name)
            pair = ImagePair(ir_path, vi_path)
            ir, vi = pair.ir_t, pair.vi_t
            start = time.time()
            ir, vi = [ir.half(), vi.half()] if self.half else [ir, vi]
            ir, vi = ir.to(self.device), vi.to(self.device)
            torch.cuda.synchronize()
            fus = self.net(ir.unsqueeze(0), vi.unsqueeze(0)).clip(0., 1.)
            torch.cuda.synchronize()
            end = time.time()
            time_list.append(end - start)
            pair.save_fus(dst / ir_path.name, fus, color)
        print(dst / ir_path.name)