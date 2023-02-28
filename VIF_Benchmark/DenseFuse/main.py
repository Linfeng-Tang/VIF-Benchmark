# Demo - train the DenseFuse network & use it to generate an image

from __future__ import print_function

import time

# from train_recons import train_recons
from generate import generate
from utils import list_images
import os
import time
import pandas as pd
from openpyxl import load_workbook
os.environ["CUDA_VISIBLE_DEVICES"] = "1"

# True for training phase
IS_TRAINING = False
# True for video sequences(frames)
IS_VIDEO = False
# True for RGB images
is_RGB = False

BATCH_SIZE = 2
EPOCHES = 4

SSIM_WEIGHTS = [1, 10, 100, 1000]
MODEL_SAVE_PATHS = [
    './models/densefuse_gray/densefuse_model_bs2_epoch4_all_weight_1e0.ckpt',
    './models/densefuse_gray/densefuse_model_bs2_epoch4_all_weight_1e1.ckpt',
    './models/densefuse_gray/densefuse_model_bs2_epoch4_all_weight_1e2.ckpt',
    './models/densefuse_gray/densefuse_model_bs2_epoch4_all_weight_1e3.ckpt',
]

# MODEL_SAVE_PATH = './models/deepfuse_dense_model_bs4_epoch2_relu_pLoss_noconv_test.ckpt'
# model_pre_path  = './models/deepfuse_dense_model_bs2_epoch2_relu_pLoss_noconv_NEW.ckpt'

# In testing process, 'model_pre_path' is set to None
# The "model_pre_path" in "main.py" is just a pre-train model and not necessary for training and testing. 
# It is set as None when you want to train your own model. 
# If you already train a model, you can set it as your model for initialize weights.
model_pre_path = None

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def writexls(save_name, method_name, time_list, sheet_name, i):
    df = pd.DataFrame({method_name: time_list})
    append_df_to_excel(save_name, df, sheet_name=sheet_name, index=False, startrow=0, startcol=i)

def main():
    time_list = []
    method_name = 'DenseFuse'
    time_save_name = '/data1/timer/Method-Comparison/Time_Statistics_all.xlsx'
    sheet_name = 'RoadScene'
    # ir_path = r'/data1/timer/Method-Comparison/New_Test_ir'
    # vi_path = r'/data1/timer/Method-Comparison/New_Test_vi'

    # ir_path = r'/data1/timer/Method-Comparison/TNO_ir_select'
    # vi_path = r'/data1/timer/Method-Comparison/TNO_vi_select'

    # ir_path = r'/data1/timer/Method-Comparison/Road_ir_select'
    # vi_path = r'/data1/timer/Method-Comparison/Road_vi_select'

    ir_path = r'/data1/timer/Method-Comparison/Dataset/RoadScene_IR'
    vi_path = r'/data1/timer/Method-Comparison/Dataset/RoadScene_VI' 
    fused_path = r'DenseFuse_Results_RoadScene_all'
    # fused_path = r'DenseFuse_Results_TNO'
    if IS_TRAINING:

        original_imgs_path = list_images('./original/')
        validatioin_imgs_path = list_images('./validation1/')

        for ssim_weight, model_save_path in zip(SSIM_WEIGHTS, MODEL_SAVE_PATHS):
            print('\nBegin to train the network ...\n')
            train_recons(original_imgs_path, validatioin_imgs_path, model_save_path, model_pre_path, ssim_weight,
                         EPOCHES, BATCH_SIZE, debug=True)

            print('\nSuccessfully! Done training...\n')
    else:
        if IS_VIDEO:
            ssim_weight = SSIM_WEIGHTS[0]
            model_path = MODEL_SAVE_PATHS[0]

            IR_path = list_images('video/1_IR/')
            VIS_path = list_images('video/1_VIS/')
            output_save_path = 'video/fused' + str(ssim_weight) + '/'
            generate(IR_path, VIS_path, model_path, model_pre_path,
                     ssim_weight, 0, IS_VIDEO, 'addition', output_path=output_save_path)
        else:
            ssim_weight = SSIM_WEIGHTS[2]
            model_path = MODEL_SAVE_PATHS[2]

            fused_path = os.path.join(os.getcwd(), fused_path)
            if not os.path.exists(fused_path):
                os.makedirs(fused_path)
            filelist = os.listdir(ir_path)
            filelist.sort(key=lambda x: int(x[0:-4]))
            for item in filelist:
                index = 0
                if item.endswith('.bmp') or item.endswith('.tif'):
                    num = int(item.split('.')[0])
                    print(num)
                    ir_image_name = os.path.join(os.path.abspath(ir_path), item)
                    vi_image_name = os.path.join(os.path.abspath(vi_path), item)
                    fused_image_name = os.path.join(os.path.abspath(fused_path), item)

                    # choose fusion layer
                    fusion_type = 'addition'
                    # fusion_type = 'l1'
                    # for ssim_weight, model_path in zip(SSIM_WEIGHTS, MODEL_SAVE_PATHS):
                    # 	output_save_path = 'outputs'
                    #
                    # 	generate(infrared, visible, model_path, model_pre_path,
                    # 	         ssim_weight, index, IS_VIDEO, is_RGB, type = fusion_type, output_path = output_save_path)

                    output_save_path = 'outputs'
                    start = time.time()
                    generate(ir_image_name, vi_image_name, model_path, model_pre_path,
                             ssim_weight, index, IS_VIDEO, is_RGB, type=fusion_type, output_path=fused_path, name=item)
                    end =time.time()
                    time_list.append(end - start)
            # i = 0
            # writexls(time_save_name, method_name, time_list, sheet_name, i)


if __name__ == '__main__':
    main()
